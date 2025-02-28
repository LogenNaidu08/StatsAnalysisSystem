import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
import openpyxl.cell._writer  
import matplotlib.pyplot as plt
from adjustText import adjust_text  # Install via `pip install adjustText`
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Image, Spacer, PageBreak
from fpdf import FPDF



class StatsCalculatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Data Analyzer")
        
        self.file_paths = {
            'file1': {'path': None, 'label': None},
            'file2': {'path': None, 'label': None},
            'file3': {'path': None, 'label': None}
        }
        
        self.start_row = tk.IntVar(value=6)  # Default start row is 4
        
        self.create_widgets()
        
    def create_widgets(self):
        main_frame = tk.Frame(self.root, padx=20, pady=20)
        main_frame.pack()
        
        # File upload section
        upload_frame = tk.Frame(main_frame)
        upload_frame.pack(pady=10)
        
        # Individual file upload buttons
        for i in range(1, 4):
            row_frame = tk.Frame(upload_frame)
            row_frame.pack(fill=tk.X, pady=5)
            
            btn = tk.Button(
                row_frame,
                text=f"Select File {i}",
                command=lambda idx=i: self.load_file(idx),
                width=12
            )
            btn.pack(side=tk.LEFT)
            
            label = tk.Label(row_frame, text="No file selected", width=40, anchor='w')
            label.pack(side=tk.LEFT, padx=10)
            self.file_paths[f'file{i}']['label'] = label

        # Row selection section
        row_frame = tk.Frame(main_frame)
        row_frame.pack(pady=10)
        
        tk.Label(row_frame, text="Start reading data from row (1-based):").pack(side=tk.LEFT, padx=5)
        tk.Entry(row_frame, textvariable=self.start_row, width=5).pack(side=tk.LEFT)

        # Process section
        process_frame = tk.Frame(main_frame)
        process_frame.pack(pady=15)
        
        self.status_label = tk.Label(process_frame, text="", fg="gray")
        self.status_label.pack(pady=5)
        
        tk.Button(
            process_frame,
            text="Calculate Statistics",
            command=self.process_files,
            bg="#4CAF50",
            fg="white"
        ).pack(pady=10)

    def load_file(self, file_number):
        file_path = filedialog.askopenfilename(
            title=f"Select File {file_number}",
            filetypes=[("Excel Files", "*.xls"), ("All Files", "*.*")]
        )
        
        if file_path:
            key = f'file{file_number}'
            self.file_paths[key]['path'] = file_path
            self.file_paths[key]['label'].config(
                text=os.path.basename(file_path),
                fg="green"
            )
            
    def process_files(self):
        # Validate inputs
        missing_files = [f"File {i}" for i in range(1, 4) 
                    if not self.file_paths[f'file{i}']['path']]
        if missing_files:
            messagebox.showerror("Missing Files", f"Please select: {', '.join(missing_files)}")
            return

        try:
            output_dir = filedialog.askdirectory(title="Select Save Location")
            if not output_dir:
                return
                
            stats = self.calculate_stats(
                file_list=[
                    self.file_paths['file1']['path'],
                    self.file_paths['file2']['path'],
                    self.file_paths['file3']['path']
                ],
                start_row=self.start_row.get() - 1  # Convert to 0-based index
            )
            
            # Save results to Excel
            output_path = os.path.join(output_dir, "results.xlsx")
            stats.to_excel(output_path, index=False, header=False)  # Save as Excel file
            
            # Format the Excel file
            self.format_excel(output_path)
            
            self.status_label.config(
                text=f"Files saved successfully to:\n{output_dir}",
                fg="darkgreen"
            )
            pdf_path = os.path.join(output_dir, "results.pdf")
            self.create_pdf(pdf_path, output_path, ["mean.png", "STD.png", "CV.png"])

            self.status_label.config(
            text=f"Files saved successfully to:\n{output_dir}",
            fg="darkgreen"
        )
            messagebox.showinfo("Success", "Processing completed!")
            
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.status_label.config(text="Processing failed", fg="red")
            
    def create_pdf(self, pdf_path, excel_path, image_paths=None):
        doc = SimpleDocTemplate(pdf_path, pagesize=landscape(letter))
        elements = []
        styles = getSampleStyleSheet()

        # Add Title
        title = Paragraph("<b>Statistical Analysis Report</b>", styles["Title"])
        elements.append(title)
        elements.append(Spacer(1, 12))

        # Load the Excel file
        wb = load_workbook(excel_path)
        ws = wb.active

        # Extract table data from Excel
        table_data = []
        merge_commands = []  # Store merging instructions

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            new_row = []
            last_value = None
            merge_start_col = None

            for j, cell in enumerate(row):
                if (cell is None or cell == "") and ((i!=1 and j!=7) or (i!=2 and j!=7) and (i!=3 and j!=7) and (i!=4 and j!=7)):  # Empty cell detected
                    if last_value is not None and merge_start_col is None:
                        merge_start_col = j - 1  # Start merging from previous cell
                    new_row.append("")  # Add an empty placeholder
                else:
                    if (merge_start_col is not None):
                        # Merge from merge_start_col to the previous column
                        merge_commands.append(('SPAN', (merge_start_col, i), (j - 1, i)))
                        merge_start_col = None  # Reset merge tracking
                    new_row.append(str(cell))  # Convert data to string
                    last_value = cell  # Store the last non-empty value
            
            # If a merge was started but not completed, finalize it
            if merge_start_col is not None:
                merge_commands.append(('SPAN', (merge_start_col, i), (len(row) - 1, i)))

            table_data.append(new_row)

        # Create a formatted table
        table = Table(table_data)

        # Define table styles
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
        ])

        # Apply merging instructions correctly
        for merge_cmd in merge_commands:
            table_style.add(*merge_cmd)  # Correct unpacking format

        table.setStyle(table_style)

        elements.append(table)
        elements.append(Spacer(1, 20))

        # Add Plots if Available
        if image_paths:
            for img_path in image_paths:
                if os.path.exists(img_path):
                    elements.append(Image(img_path, width=500, height=300))
                    elements.append(Spacer(1, 15))

        # Save the PDF
        doc.build(elements)
        messagebox.showinfo("PDF Saved", f"PDF saved successfully: {pdf_path}")


    def format_excel(self, file_path):
        
        # Load the workbook
        wb = load_workbook(file_path)
        ws = wb.active

        # Define border style
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # Apply borders to all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Format the headers (Row 1)
        header_row = ws[1]
        for cell in header_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        #merge cells for H1 to J5
        ws.merge_cells("H1:J5")
        
        # Merge cells for INTERVAL and times
        ws.merge_cells("B1:C1")  # Merge 9AM cells
        ws.merge_cells("D1:E1")  # Merge 12PM cells
        ws.merge_cells("F1:G1")  # Merge 3PM cells
        #ws.merge_cells("H1:J1")  # Merge Statistical Analysis cells

        
        # Format the N/C values (Rows 2-4)
        NC_rows = ws[2:4]
        for cell in NC_rows:
            for c in cell:
                if c == cell[0]:
                    c.font = Font(bold=True)
                else:
                    c.font = Font(bold=False)
                c.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B2:C2")  # Merge NC1 cells
        ws.merge_cells("D2:E2")  # Merge NC1 cells
        ws.merge_cells("F2:G2")  # Merge NC1 cells
        #ws.merge_cells("H2:J2")  # Merge Statistical Analysis cells
        ws.merge_cells("B3:C3")  # Merge NC2 cells  
        ws.merge_cells("D3:E3")  # Merge NC2 cells
        #ws.merge_cells("H3:J3")  # Merge NC2 cells
        ws.merge_cells("F3:G3")  # Merge NC2 cells 
        ws.merge_cells("B4:C4")  # Merge NC3 cells  
        ws.merge_cells("D4:E4")  # Merge NC3 cells
        ws.merge_cells("F4:G4")  # Merge NC3 cells 
        #ws.merge_cells("H4:J4")  # Merge Statistical Analysis cells

        # Format the CUT-OFF values (Row 5)
        cutoff_row = ws[5]
        for cell in cutoff_row:
            if cell == cutoff_row[0]:
                cell.font = Font(bold=True)
            else:
                cell.font = Font(bold=False)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("B5:C5")  # Merge CUT-OFF1 cells
        ws.merge_cells("D5:E5")  # Merge CUT-OFF2 cells
        ws.merge_cells("F5:G5")  # Merge CUT-OFF2 cells 
        #ws.merge_cells("H5:J5")  # Merge Statistical Analysis cells

        sub_header_row = ws[6]
        for cell in sub_header_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Format the SAMPLES (Row 7 -26)
        samples_row = ws[7:11]
        for cell in samples_row:
            for c in cell:
                if c == cell[0]:
                    c.font = Font(bold=True)
                else:
                    c.font = Font(bold=False)
                c.alignment = Alignment(horizontal="center", vertical="center")

         # Add special border for summary section
        summary_start = 12  # Adjust based on your row count
        for row in ws.iter_rows(min_row=summary_start):
            for cell in row:
                cell.border = thin_border
        # Adjust column widths
        column_widths = [15, 10, 10, 10, 10, 10, 10]  # Adjust as needed
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Save the workbook with formatting
        wb.save(file_path)
    
    def calculate_stats(self, file_list, start_row):
        data = []
        ncs = []  # To store N/C values for each file

        # Define the rows for N/C values (0-based indices)
        nc_rows = [5, 7, 8]  # First 3 rows
        nc_columns = [1]  # Columns you want to select

        for file in file_list:
            df = pd.read_excel(file, header=None)
            
            # Extract N/C values (first 3 rows, all columns)
            nc_values = df.iloc[nc_rows, nc_columns].values.flatten()  # Flatten to 1D array
            ncs.append(nc_values)
            
            # Extract sample data starting from the user-specified row
            sample_data = df.iloc[start_row:, :].apply(pd.to_numeric, errors='coerce')
            data.append(sample_data)

        # Calculate cutoff values (average of N/C values + 0.1 for each file)
        ncs_array = np.array(ncs)  # Shape: (3 files, N N/C values)
        cutoff_values = np.round(np.nanmean(ncs_array, axis=1) + 0.1, 3)  # Shape: (3,)
        print(ncs_array)
        print(ncs_array.shape)

        # Process sample data
        data_array = np.array([df.values for df in data])  # Shape: (3 files, rows, cols)
        print(data_array)
        print(data_array.shape)
        
        # Calculate mean, std, and CV across files
        mean_data = np.round(np.nanmean(data_array, axis=0),3)
        print(mean_data)
        std_data = np.round(np.nanstd(data_array, axis=0),3)
        print(std_data)
        with np.errstate(divide='ignore', invalid='ignore'):
            cv_data = np.round(np.where(mean_data != 0, (std_data / mean_data) * 100, np.nan),3)
        print(cv_data)

        # Determine Positive/Negative results
        results = []
        for file_idx in range(data_array.shape[0]):
            file_results = np.where(data_array[file_idx] > cutoff_values[file_idx], "Positive", "Negative")
            results.append(file_results)

        # Create output DataFrame
        output_rows = []

        # Add headers
        output_rows.append(["INTERVAL", "9AM", "", "12PM", "", "3PM", "", "Statistical Analysis"])

        # Add N/C values
        for nc_idx in range(ncs_array.shape[1]):
                output_rows.append([f"NC{nc_idx + 1}", ncs_array[0, nc_idx], "", ncs_array[1, nc_idx], "", ncs_array[2, nc_idx]])

        # Add cutoff values
        output_rows.append(["CUT-OFF", cutoff_values[0], "", cutoff_values[1], "", cutoff_values[2]])

        # Add sample data
        output_rows.append(["SAMPLES", "O/D", "Results", "O/D", "Results", "O/D", "Results", "Mean", "Std.Dev", "CV (%)"])
        #idx2 = 5
        #idx1 = 3
        #idx3 = 7
        for sample_idx in range(5):
            row = [f"SAMPLE {sample_idx + 1}"]
            if sample_idx < 3:
                row.extend([data_array[0, sample_idx+6, 1], results[0][sample_idx+6, 1]])
                row.extend([data_array[1, sample_idx+6, 1], results[1][sample_idx+6, 1]])
                row.extend([data_array[2, sample_idx+6, 1], results[2][sample_idx+6, 1]])
                if sample_idx + 5 == 5:
                    row.extend([mean_data[6, 1], std_data[6, 1], cv_data[6, 1]])
                elif sample_idx + 5 == 6:
                    row.extend([mean_data[7, 1], std_data[7, 1], cv_data[7, 1]])
                elif sample_idx + 5 == 7:
                    row.extend([mean_data[8, 1], std_data[8, 1], cv_data[8, 1]])

            elif sample_idx >= 3 and sample_idx <= 5:
                if sample_idx !=4:
                    row.extend([data_array[0, sample_idx-3, 3], results[0][sample_idx-3, 3]])
                    row.extend([data_array[1, sample_idx-3, 3], results[1][sample_idx-3, 3]])
                    row.extend([data_array[2, sample_idx-3, 3], results[2][sample_idx-3, 3]])
                    row.extend([mean_data[sample_idx-3, 3], std_data[sample_idx-3, 3], cv_data[sample_idx-3, 3]])
                elif sample_idx == 4:
                    row.extend([data_array[0, sample_idx-2, 3], results[0][sample_idx-2, 3]])
                    row.extend([data_array[1, sample_idx-2, 3], results[1][sample_idx-2, 3]])
                    row.extend([data_array[2, sample_idx-2, 3], results[2][sample_idx-2, 3]])
                    row.extend([mean_data[sample_idx-2, 3], std_data[sample_idx-2, 3], cv_data[sample_idx-2, 3]])
            """
            elif sample_idx >= 11 and sample_idx <= 18:
                if sample_idx != 12:
                    row.extend([data_array[0, sample_idx-11, 5], results[0][sample_idx-11, 5]])
                    row.extend([data_array[1, sample_idx-11, 5], results[1][sample_idx-11, 5]])
                    row.extend([data_array[2, sample_idx-11, 5], results[2][sample_idx-11, 5]])
                    row.extend([mean_data[sample_idx-11, 5], std_data[sample_idx-11, 5], cv_data[sample_idx-11, 5]])
                elif sample_idx == 12:
                    row.extend([data_array[0, sample_idx-10, 5], results[0][sample_idx-10, 5]])
                    row.extend([data_array[1, sample_idx-10, 5], results[1][sample_idx-10, 5]])
                    row.extend([data_array[2, sample_idx-10, 5], results[2][sample_idx-10, 5]])
                    row.extend([mean_data[sample_idx-10, 5], std_data[sample_idx-10, 5], cv_data[sample_idx-10, 5]])

            elif sample_idx == 19 :
                    row.extend([data_array[0, sample_idx-19, 7], results[0][sample_idx-19, 7]])
                    row.extend([data_array[1, sample_idx-19, 7], results[1][sample_idx-19, 7]])
                    row.extend([data_array[2, sample_idx-19, 7], results[2][sample_idx-19, 7]])
                    row.extend([mean_data[sample_idx-19, 7], std_data[sample_idx-19, 7], cv_data[sample_idx-19, 7]])"""
            output_rows.append(row)

        # Add summary
        """summary_text = [
            "SUMMARY:",
            "Based on the result above, we can conclude that Meril Merilisa Malaria PAN Ag (ELISA kit for Malaria PAN Ag (pLDH)) passed according to manufacturing or principal claims. Below is the claim from the manufacturer or principle. Please refer to the insert/IFU we attached together with report as references.",
            "Precision claim: CV < 3.0%",
            "Sensitivity claim: 100%",
            "Specificity claim: 100%",
            "The standard deviation and CV value is to measure the amount of variation or dispersion of a set of value. The smaller the value, the better the test result."
        ]
        for text in summary_text:
            output_rows.append([text])"""
        
        # Initialize lists to store values
        sample_labels = []
        mean_values = []
        std_values = []
        cv_values = []

        # Extract data from the loop
        #idx = 3
        #idx1 = 3
        for sample_idx in range(5):
            sample_labels.append(f"SAMPLE {sample_idx + 1}")
            
            if sample_idx < 3:
                mean_values.append(mean_data[sample_idx+6, 1])
                std_values.append(std_data[sample_idx+6, 1])
                cv_values.append(cv_data[sample_idx+6, 1])
            elif sample_idx >= 3 and sample_idx <= 4:
                if sample_idx != 4:
                    mean_values.append(mean_data[sample_idx-3, 3])
                    std_values.append(std_data[sample_idx-3, 3])
                    cv_values.append(cv_data[sample_idx-3, 3])
                elif sample_idx == 4:
                    mean_values.append(mean_data[sample_idx-2, 3])
                    std_values.append(std_data[sample_idx-2, 3])
                    cv_values.append(cv_data[sample_idx-2, 3])
            """
            elif sample_idx >= 11 and sample_idx <= 18:
                if sample_idx != 12:
                    mean_values.append(mean_data[sample_idx-11, 5])
                    std_values.append(std_data[sample_idx-11, 5])
                    cv_values.append(cv_data[sample_idx-11, 5])
                elif sample_idx == 12:
                    mean_values.append(mean_data[sample_idx-10, 5])
                    std_values.append(std_data[sample_idx-10, 5])
                    cv_values.append(cv_data[sample_idx-10, 5])
            elif sample_idx == 19:
                mean_values.append(mean_data[sample_idx-19, 7])
                std_values.append(std_data[sample_idx-19, 7])
                cv_values.append(cv_data[sample_idx-19, 7])
            """
        #plotting the graph for mean
        plt.figure(figsize=(10, 5))
        plt.plot(sample_labels, mean_values, color='green', marker='o', linestyle='dashed', linewidth=2, markersize=12, markerfacecolor='red',  label='Data')
        # Show values on the points
        texts_mean = []
        for i, txt in enumerate(mean_values):
            texts_mean.append(plt.text(sample_labels[i], mean_values[i], str(txt), fontsize=12))
        # Adjust the position of the text
        adjust_text(texts_mean, only_move={'points': 'xy', 'text': 'xy'}, arrowprops=dict(arrowstyle="-", color='gray'))
        plt.xlabel("Samples")
        plt.ylabel("Mean")
        plt.title("Comparison of Mean Across Samples")
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.savefig("mean.png")
        plt.close()

        #plotting the graph for std
        plt.figure(figsize=(10, 5))
        plt.plot(sample_labels, std_values, color='green', marker='o', linestyle='dashed', linewidth=2, markersize=12, markerfacecolor='red',  label='Data')
        # Show values on the points
        texts_std = []
        for i, txt in enumerate(std_values):
            texts_std.append(plt.text(sample_labels[i], std_values[i], str(txt), fontsize=12))
        # Adjust the position of the text
        adjust_text(texts_std, only_move={'points': 'xy', 'text': 'xy'}, arrowprops=dict(arrowstyle="-", color='gray'))
        plt.xlabel("Samples")
        plt.ylabel("STD")
        plt.title("Comparison of STD Across Samples")
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.savefig("STD.png")
        plt.close()

        #plotting the graph for cv
        plt.figure(figsize=(10, 5))
        plt.plot(sample_labels, cv_values, color='green', marker='o', linestyle='dashed', linewidth=2, markersize=12, markerfacecolor='red',  label='Data')
        # Show values on the points
        #texts_cv = []
        #for i, txt in enumerate(cv_values):
            #texts_cv.append(plt.text(sample_labels[i], cv_values[i], str(txt), fontsize=12))
        # Adjust the position of the text
        #adjust_text(texts_cv, only_move={'points': 'xy', 'text': 'xy'}, arrowprops=dict(arrowstyle="-", color='gray'))
        plt.xlabel("Samples")
        plt.ylabel("CV")
        plt.title("Comparison of CV Across Samples")
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.savefig("CV.png")
        plt.close()

        # Convert to DataFrame
        output_df = pd.DataFrame(output_rows)
        return output_df

if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("600x400")
    StatsCalculatorApp(root)
    root.mainloop()